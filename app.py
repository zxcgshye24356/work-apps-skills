# -*- coding: utf-8 -*-
"""
百度收录效果批量检测工具（双品牌词版）v2
框架：Streamlit + Playwright + pandas/openpyxl + 大模型API(OpenAI格式)
更新：
  - 稳定性增强：固定 UA、百度验证码自动重试、多次运行取趋势
  - 汇总看板：归类/好差/品牌对比图表
  - 链接可点击：结果表 + 链接明细表（HYPERLINK 可点击）
  - 历史记录：每次运行自动存档，界面可查看/下载/对比
"""

import io
import json
import os
import random
import re
import time
import urllib.parse

import pandas as pd
import streamlit as st

# ---------- 固定品牌词（按需求文档）----------
BRAND_A = "陕西新研博美"
BRAND_B = "西安凯新生物"

# ---------- 反爬：固定 User-Agent（去掉随机，提升结果稳定性）----------
FIXED_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# 用于抹掉自动化痕迹的初始化脚本
STEALTH_JS = """
() => {
  Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
  try {
    Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3] });
  } catch (e) {}
  try {
    Object.defineProperty(navigator, 'languages', { get: () => ['zh-CN', 'zh', 'en'] });
  } catch (e) {}
}
"""

# 输出列（含稳定性增强新增列）
OUT_COLUMNS = [
    "产品名称", "品牌词归类", "品牌词出现详情", "收录效果状态",
    "数据新鲜度", "趋势结论", "文章链接列表", "备注",
]

HISTORY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history")


# ================= 核心检测逻辑 =================

def search_baidu(context, query, max_retry=3):
    """在百度搜索第一页，返回 (links列表, 是否被安全验证拦截)。
    若触发百度安全验证，会等待后自动重试；全部失败才返回 blocked=True。
    """
    blocked = False
    for attempt in range(max_retry):
        page = context.new_page()
        try:
            url = "https://www.baidu.com/s?wd=" + urllib.parse.quote(query)
            page.goto(url, timeout=20000, wait_until="domcontentloaded")
            try:
                page.wait_for_selector("#content_left", timeout=10000)
            except Exception:
                pass
            time.sleep(1.5)  # 等待懒加载结果，让第一页链接抓得更全
            content = page.content()
            if ("百度安全验证" in content) or ("wappass" in page.url):
                blocked = True
                try:
                    page.close()
                except Exception:
                    pass
                time.sleep(3)  # 触发验证时停顿后重试
                continue
            links = page.evaluate(
                """
                () => {
                    const out = [];
                    const blocks = document.querySelectorAll('#content_left .c-container, #content_left .result');
                    blocks.forEach(b => {
                        const a = b.querySelector('h3 a') || b.querySelector('a');
                        if (a && a.href) out.push(a.href);
                    });
                    return Array.from(new Set(out));
                }
                """
            )
            try:
                page.close()
            except Exception:
                pass
            return (links or []), False
        except Exception:
            try:
                page.close()
            except Exception:
                pass
            return [], False
    return [], blocked


def split_queries(name):
    """将单元格内的多个名称/CAS 拆分为独立检索词。
    分隔符：空格、/、\\、，、,、；、;、|、｜。
    不含连字符（CAS 号 / DSPE-PEG 等需整体检索）。
    """
    s = (name or "").strip()
    if not s:
        return [s]
    if is_cas_like(s):
        return [s]
    parts = re.split(r"[ \t/\\，,；;|｜]+", s)
    parts = [p.strip() for p in parts if p and p.strip()]
    if not parts:
        return [s]
    if len(parts) == 1:  # 没有分隔符，整体检索
        return [s]
    return parts


def search_product(context, name):
    """拆分产品名为多个检索词，分别搜百度第一页，链接去重合并。
    返回 (links列表, 是否全部被拦截, 检索词列表)。
    """
    queries = split_queries(name)
    merged, seen = [], set()
    blocked_all = True
    for q in queries:
        links, blocked = search_baidu(context, q)
        if not blocked:
            blocked_all = False
        for u in (links or []):
            if u not in seen:
                seen.add(u)
                merged.append(u)
    return merged, blocked_all, queries


def extract_text(page):
    """提取页面正文文本（innerText），失败返回空。"""
    try:
        text = page.evaluate("() => (document.body ? document.body.innerText : '')")
        return (text or "").strip()
    except Exception:
        return ""


def detect_local(text, brand_a, brand_b):
    """本地精确匹配：品牌词完整出现即算提及。"""
    return (brand_a in text, brand_b in text)


def is_cas_like(name):
    """识别 CAS 号格式（如 1159408-67-9），避免被误当联系方式处理/过滤。"""
    return bool(re.match(r"^\d{2,7}-\d{1,2}-\d{1}$", (name or "").strip()))


def detect_llm(text, brand_a, brand_b, client, model):
    """调用大模型API判断文本是否提及两个品牌词。返回 (a提及, b提及)。"""
    snippet = text[:3000]  # Token 控制：超过 5000 字只取前 3000 字
    prompt = (
        "你是一个文本分析助手。请判断下列文本是否提及了两个品牌。\n"
        f"品牌A名称：{brand_a}\n"
        f"品牌B名称：{brand_b}\n"
        "规则：\n"
        "- 文本中完整出现品牌A名称即认为提及了A。\n"
        "- 文本中完整出现品牌B名称即认为提及了B。\n"
        "- 一篇文本可能同时提及两个品牌，也可能都不提及。\n"
        "只输出JSON，不要输出其他任何内容，格式：{\"A\": true/false, \"B\": true/false}\n"
        "文本内容如下：\n" + snippet
    )
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )
        raw = resp.choices[0].message.content
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if not m:
            return (False, False)
        obj = json.loads(m.group(0))  # JSON 的 true/false 由 json 正确解析
        return (bool(obj.get("A", False)), bool(obj.get("B", False)))
    except Exception:
        # 大模型失败时回退到本地匹配，保证流程不中断
        return detect_local(text, brand_a, brand_b)


def analyze_link(context, href, brand_a, brand_b, detect_fn, detect_kwargs, timeout=15000, extra_brand=None):
    """打开单篇链接，提取正文并检测品牌词。成功返回 dict，无法解析返回 None。
    extra_brand：当指定品牌不是 A/B 时的自定义品牌词，额外检测其是否出现。
    """
    page = context.new_page()
    try:
        page.goto(href, timeout=timeout, wait_until="domcontentloaded")
        url = page.url
        text = extract_text(page)
        if len(text) < 20:  # 视频/PDF/反爬页等无法解析
            return None
        a, b = detect_fn(text, brand_a, brand_b, **detect_kwargs)
        lock = False
        if extra_brand:
            la, _ = detect_fn(text, extra_brand, extra_brand, **detect_kwargs)
            lock = la
        return {"url": url, "a": a, "b": b, "lock": lock}
    except Exception:
        return None
    finally:
        try:
            page.close()
        except Exception:
            pass


def process_product(context, name, brand_a, brand_b, detect_fn, detect_kwargs, brand_lock=None):
    """处理单个产品（单次运行），返回结果字典。
    brand_lock 不为空时（用户已在品牌列指定）：
      - 品牌词归类直接锁定为 brand_lock；
      - 算法1：只统计「该指定品牌词」被提及的篇数，用于判断收录效果；
      - 不再比较两个品牌谁多。
    """
    cas = is_cas_like(name) or any(is_cas_like(q) for q in split_queries(name))
    cas_note = "（CAS号检索）" if cas else ""
    links, blocked, queries = search_product(context, name)

    if blocked:
        return {
            "产品名称": name,
            "品牌词归类": "",
            "品牌词出现详情": "",
            "收录效果状态": "",
            "文章链接列表": "",
            "备注": "触发百度安全验证，请稍后重试" + cas_note,
        }

    if not links:
        return {
            "产品名称": name,
            "品牌词归类": "",
            "品牌词出现详情": "",
            "收录效果状态": "",
            "文章链接列表": "",
            "备注": "未搜索到" + cas_note,
        }

    # 指定品牌且非 A/B 时，需额外检测该自定义词
    extra = brand_lock if (brand_lock and brand_lock not in (brand_a, brand_b)) else None

    a_count = b_count = lock_count = 0
    real_urls = []
    failed = 0

    for href in links:
        res = analyze_link(context, href, brand_a, brand_b, detect_fn, detect_kwargs, extra_brand=extra)
        if res is None:
            failed += 1
            continue
        real_urls.append(res["url"])
        if res["a"]:
            a_count += 1  # 一篇文章对单一品牌只计一次
        if res["b"]:
            b_count += 1
        if brand_lock:
            if brand_lock == brand_a:
                hit = res["a"]
            elif brand_lock == brand_b:
                hit = res["b"]
            else:
                hit = res["lock"]
            if hit:
                lock_count += 1

    if brand_lock:
        cls = brand_lock
        count = lock_count
        detail = f"{brand_lock}:{count}篇"
        status = "好" if count >= 5 else "差"
        lock_note = "用户指定品牌，已锁定归类"
    else:
        if a_count > b_count:
            cls = brand_a
        elif b_count > a_count:
            cls = brand_b
        else:
            cls = "持平"
        count = a_count + b_count
        detail = f"{brand_a}:{a_count}篇，{brand_b}:{b_count}篇"
        status = "好" if count >= 5 else "差"
        lock_note = ""

    note_parts = []
    if lock_note:
        note_parts.append(lock_note)
    if cas:
        note_parts.append("CAS号检索")
    if len(queries) > 1:
        note_parts.append(f"拆分为{len(queries)}个检索词")
    if failed:
        note_parts.append(f"{failed}个链接无法解析")
    note = "｜".join(note_parts)

    return {
        "产品名称": name,
        "品牌词归类": cls,
        "品牌词出现详情": detail,
        "收录效果状态": status,
        "文章链接列表": "，".join(real_urls),
        "备注": note,
    }


def _parse_counts(detail):
    """从 'A:x篇，B:y篇' 解析出 (a, b) 整数。"""
    nums = re.findall(r"(\d+)篇", detail or "")
    a = int(nums[0]) if len(nums) >= 1 else 0
    b = int(nums[1]) if len(nums) >= 2 else 0
    return a, b


def run_batch(context, names, brand_a, brand_b, detect_fn, detect_kwargs, runs=1, brand_map=None):
    """对每个产品跑 runs 次，聚合为最终行 + 各产品链接集合。
    brand_map: {产品名: 指定品牌} 或 None（自动判断模式）。
    返回 (final_rows, per_product_links)。
    """
    per_run = []
    for _ in range(runs):
        run_rows = []
        for name in names:
            time.sleep(random.uniform(0.8, 1.8))  # 适度随机延迟，规避反爬
            lock = brand_map.get(name) if brand_map else None
            if lock == "":
                lock = None
            run_rows.append(process_product(context, name, brand_a, brand_b, detect_fn, detect_kwargs, brand_lock=lock))
        per_run.append(run_rows)

    final_rows = []
    per_product_links = {}
    for i, name in enumerate(names):
        runs_rows = [per_run[r][i] for r in range(runs)]
        lock = brand_map.get(name) if brand_map else None
        if lock == "":
            lock = None

        # 链接跨运行去重合并
        merged = []
        seen = set()
        for rr in runs_rows:
            for u in (rr.get("文章链接列表") or "").split("，"):
                u = u.strip()
                if u and u not in seen:
                    seen.add(u)
                    merged.append(u)
        per_product_links[name] = merged

        blocked = sum(1 for rr in runs_rows if "安全验证" in rr.get("备注", ""))
        nosearch = sum(1 for rr in runs_rows if rr.get("品牌词归类", "") == "")
        failed = sum(1 for rr in runs_rows if "无法解析" in rr.get("备注", ""))

        if lock:
            # 模式A：只统计指定品牌篇数（算法1）
            lock_list = []
            for rr in runs_rows:
                m = re.search(r"(\d+)篇", rr.get("品牌词出现详情", ""))
                lock_list.append(int(m.group(1)) if m else 0)
            lock_avg = round(sum(lock_list) / len(lock_list)) if lock_list else 0
            cls = lock
            status = "好" if lock_avg >= 5 else "差"
            detail = f"{lock}:{lock_avg}篇"
            trend = "用户指定品牌（已锁定）"
            note = f"共运行{runs}次：{lock}平均{lock_avg}篇"
        else:
            a_list, b_list = [], []
            win_a = win_b = tie = 0
            for rr in runs_rows:
                a, b = _parse_counts(rr.get("品牌词出现详情", ""))
                a_list.append(a)
                b_list.append(b)
                if a > b:
                    win_a += 1
                elif b > a:
                    win_b += 1
                else:
                    tie += 1

            a_avg = round(sum(a_list) / len(a_list)) if a_list else 0
            b_avg = round(sum(b_list) / len(b_list)) if b_list else 0

            # 趋势结论：多数决（A赢最多且不少于持平数；否则B；否则持平）
            if win_a > win_b and win_a >= tie:
                trend = brand_a
            elif win_b > win_a and win_b >= tie:
                trend = brand_b
            else:
                trend = "持平"

            valid = [rr for rr in runs_rows if rr.get("品牌词归类", "") != ""]
            cls = trend if valid else ""

            total = a_avg + b_avg
            status = "好" if total >= 5 else "差"
            detail = f"{brand_a}:{a_avg}篇，{brand_b}:{b_avg}篇"
            note = f"共运行{runs}次：A赢{win_a}/B赢{win_b}/持平{tie}"

        if blocked:
            note += f"，{blocked}次触发验证"
        if nosearch:
            note += f"，{nosearch}次未搜索到"
        if failed:
            note += f"，{failed}次链接无法解析"

        final_rows.append({
            "产品名称": name,
            "品牌词归类": cls,
            "品牌词出现详情": detail,
            "收录效果状态": status,
            "数据新鲜度": f"跨{runs}次共抓到{len(merged)}篇" if merged else "无文章",
            "趋势结论": trend,
            "文章链接列表": "，".join(merged),
            "备注": note,
        })
    return final_rows, per_product_links


# ================= Excel 读写 =================

def load_products(uploaded_bytes):
    """读取Excel，返回 (去重后产品列表, 重复产品集合, 品牌映射)。
    若任一列名含「品牌」，进入“指定品牌模式”：
      - 品牌列 = 该列；产品列 = 含「产品」的列，否则品牌列后第一列
      - brand_map = {产品名: 品牌值}（品牌值为空字符串表示该行按自动判断）
    否则进入“自动判断模式”：brand_map = None。
    """
    df = pd.read_excel(io.BytesIO(uploaded_bytes))
    df.columns = [str(c).strip() for c in df.columns]

    brand_col = None
    for c in df.columns:
        if "品牌" in c:
            brand_col = c
            break

    if brand_col is not None:
        prod_col = None
        for c in df.columns:
            if "产品" in c and c != brand_col:
                prod_col = c
                break
        if prod_col is None:
            idx = list(df.columns).index(brand_col)
            prod_col = df.columns[idx + 1] if idx + 1 < len(df.columns) else df.columns[0]
        names_raw = df[prod_col].tolist()
        brand_raw = df[brand_col].tolist()
        brand_map = {}
    else:
        prod_col = "产品名称" if "产品名称" in df.columns else df.columns[0]
        names_raw = df[prod_col].tolist()
        brand_raw = [None] * len(names_raw)
        brand_map = None

    seen = set()
    unique = []
    duplicates = set()
    for n, b in zip(names_raw, brand_raw):
        s = str(n).strip()
        if s in ("", "nan", "None"):
            continue
        if s in seen:
            duplicates.add(s)
            continue
        seen.add(s)
        unique.append(s)
        if brand_map is not None:
            bv = str(b).strip() if str(b).strip() not in ("", "nan", "None") else ""
            brand_map[s] = bv

    return unique, duplicates, brand_map


def build_output_df(results, duplicates):
    """汇总结果，重复项单独成行标注。"""
    rows = list(results)
    for d in sorted(duplicates):
        row = {c: "" for c in OUT_COLUMNS}
        row["产品名称"] = d
        row["备注"] = "产品名称重复"
        rows.append(row)
    return pd.DataFrame(rows, columns=OUT_COLUMNS)


def export_workbook(final_rows, per_product_links):
    """导出双表 Excel：结果表 + 链接明细表（链接可点击）。返回 bytes。"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "结果"
    ws.append(OUT_COLUMNS)
    for row in final_rows:
        ws.append([row.get(c, "") for c in OUT_COLUMNS])

    ws2 = wb.create_sheet("链接明细")
    ws2.append(["产品名称", "序号", "文章链接"])
    for name, links in per_product_links.items():
        for i, u in enumerate(links, 1):
            # Excel 超链接公式，点击即可打开
            ws2.append([name, i, f'=HYPERLINK("{u}", "打开")'])

    for sheet in (ws, ws2):
        for col in sheet.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[letter].width = min(max(width + 2, 12), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_sample_xlsx():
    sample = pd.DataFrame({"产品名称": ["聚乙二醇", "DSPE-PEG", "巯基聚乙二醇", "生物素聚乙二醇"]})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sample.to_excel(writer, index=False)
    return buf.getvalue()


def make_sample_branded_xlsx():
    """含品牌列的示例：A列品牌，B列产品名（含一个 CAS 号示例）。"""
    sample = pd.DataFrame({
        "品牌": ["陕西新研博美", "西安凯新生物", "陕西新研博美", "西安凯新生物"],
        "产品名称": ["聚乙二醇", "DSPE-PEG", "1159408-67-9", "生物素聚乙二醇"],
    })
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sample.to_excel(writer, index=False)
    return buf.getvalue()


# ================= 历史记录 =================

def save_history(out_df, meta):
    """把本次运行结果存到 history/ 目录，返回存档文件名（不含后缀）。"""
    os.makedirs(HISTORY_DIR, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(HISTORY_DIR, f"run_{stamp}.csv")
    out_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    meta_path = os.path.join(HISTORY_DIR, f"run_{stamp}.json")
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    return f"run_{stamp}"


def load_history_list():
    """读取 history/ 下的运行记录元信息，按时间倒序返回列表。"""
    if not os.path.isdir(HISTORY_DIR):
        return []
    items = []
    for fn in os.listdir(HISTORY_DIR):
        if fn.endswith(".json"):
            stamp = fn[len("run_"):-len(".json")]
            try:
                with open(os.path.join(HISTORY_DIR, fn), "r", encoding="utf-8") as f:
                    meta = json.load(f)
                meta["stamp"] = stamp
                items.append(meta)
            except Exception:
                continue
    items.sort(key=lambda x: x.get("stamp", ""), reverse=True)
    return items


def read_history_csv(stamp):
    path = os.path.join(HISTORY_DIR, f"run_{stamp}.csv")
    if os.path.isfile(path):
        return pd.read_csv(path)
    return None


# ================= 汇总看板 =================

def show_dashboard(out_df, brand_a, brand_b):
    valid = out_df[out_df["品牌词归类"].notna() & (out_df["品牌词归类"].astype(str) != "")]
    if valid.empty:
        st.info("暂无有效数据进行汇总。")
        return

    st.subheader("汇总看板")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("**品牌词归类分布**")
        cls_counts = valid["品牌词归类"].value_counts()
        st.bar_chart(cls_counts)
    with c2:
        st.markdown("**收录效果（好/差）**")
        st.bar_chart(out_df["收录效果状态"].value_counts())
    with c3:
        st.markdown("**两品牌整体提及篇数**")
        a_total = b_total = 0
        for _, r in out_df.iterrows():
            a, b = _parse_counts(r.get("品牌词出现详情", ""))
            a_total += a
            b_total += b
        st.bar_chart(pd.Series({brand_a: a_total, brand_b: b_total}))

    st.caption("说明：多次运行时篇数为各次运行的平均值之和，用于横向对比两品牌整体收录强度。")


# ================= Streamlit 界面 =================

def main():
    st.set_page_config(page_title="百度收录效果批量检测工具", layout="wide")
    st.title("百度收录效果批量检测工具（双品牌词版）")
    st.caption(f"检测品牌：{BRAND_A} ｜ {BRAND_B}")

    tab_analysis, tab_history = st.tabs(["分析", "历史记录"])

    # ---------------- 分析页 ----------------
    with tab_analysis:
        st.sidebar.header("检测方式")
        mode = st.sidebar.radio(
            "品牌词检测方式",
            ["本地精确匹配（离线，推荐）", "大模型API（OpenAI格式）"],
            index=0,
            help="需求文档要求使用大模型API；本地精确匹配为离线备选，无需配置密钥也能运行。",
        )
        use_llm = mode.startswith("大模型")

        client = None
        model = ""
        if use_llm:
            base_url = st.sidebar.text_input("API Base URL", value="https://api.openai.com/v1")
            api_key = st.sidebar.text_input("API Key", type="password")
            model = st.sidebar.text_input("模型名称", value="gpt-4o-mini")
            if api_key:
                from openai import OpenAI
                client = OpenAI(base_url=base_url, api_key=api_key)
            else:
                st.sidebar.warning("未填写 API Key，将自动回退到本地精确匹配。")

        runs = st.sidebar.slider(
            "运行次数（取趋势，提升稳定性）",
            min_value=1, max_value=3, value=1,
            help="同一批产品跑多次，按多数决给出趋势结论；次数越多越稳但越慢。",
        )

        st.sidebar.markdown("---")
        st.sidebar.info(
            "首次运行需安装浏览器内核：\n```\nplaywright install chromium\n```\n"
            "并安装依赖：\n```\npip install -r requirements.txt\n```"
        )

        uploaded = st.file_uploader(
            "上传产品列表 Excel（.xlsx）。含「品牌」列→按指定品牌判断；仅「产品名称」列→自动判断品牌",
            type=["xlsx"],
        )

        if st.button("下载示例（仅产品名）"):
            st.download_button(
                "点击保存示例",
                data=make_sample_xlsx(),
                file_name="示例产品列表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        if st.button("下载示例（含品牌列）"):
            st.download_button(
                "点击保存示例",
                data=make_sample_branded_xlsx(),
                file_name="示例产品列表_含品牌列.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        if uploaded is not None:
            try:
                unique, duplicates, brand_map = load_products(uploaded.getvalue())
            except Exception as e:
                st.error(f"读取Excel失败：{e}")
                return

            if brand_map is not None:
                st.success(
                    f"识别为【指定品牌模式】共读取 {len(unique) + len(duplicates)} 条，"
                    f"去重后处理 {len(unique)} 条，重复 {len(duplicates)} 条。"
                    "（品牌列已填写的产品将锁定归类并只判断收录效果；留空的按自动判断）"
                )
                with st.expander("查看品牌映射"):
                    for n in unique:
                        b = brand_map.get(n, "")
                        st.write(f"{n} → {b if b else '（空，按自动判断）'}")
            else:
                st.success(
                    f"识别为【自动判断品牌模式】共读取 {len(unique) + len(duplicates)} 条，"
                    f"去重后处理 {len(unique)} 条，重复 {len(duplicates)} 条。"
                )
            if duplicates:
                st.warning("重复产品（不搜索，标注“产品名称重复”）：" + "、".join(sorted(duplicates)))
            with st.expander("查看去重后产品列表"):
                for i, n in enumerate(unique, 1):
                    st.write(f"{i}. {n}{'（CAS号）' if is_cas_like(n) else ''}")

            if st.button("开始分析", type="primary"):
                from playwright.sync_api import sync_playwright

                if use_llm and client is not None:
                    def detect_fn(text, a, b, **kw):
                        return detect_llm(text, a, b, client=client, model=model)
                else:
                    def detect_fn(text, a, b, **kw):
                        return detect_local(text, a, b)
                detect_kwargs = {}

                progress = st.progress(0)
                status_text = st.empty()
                all_final = []
                all_links = {}

                with st.spinner("正在启动浏览器…"):
                    with sync_playwright() as p:
                        browser = p.chromium.launch(
                            headless=True,
                            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
                        )
                        context = browser.new_context(
                            user_agent=FIXED_UA,
                            viewport={"width": 1280, "height": 800},
                        )
                        context.add_init_script(STEALTH_JS)
                        try:
                            total = len(unique)
                            for idx, name in enumerate(unique, 1):
                                status_text.info(f"正在处理 ({idx}/{total})：{name}")
                                rows, links = run_batch(
                                    context, [name], BRAND_A, BRAND_B, detect_fn, detect_kwargs, runs=runs, brand_map=brand_map
                                )
                                all_final.extend(rows)
                                all_links.update(links)
                                progress.progress(idx / total)
                        finally:
                            context.close()
                            browser.close()

                status_text.success("分析完成！")
                out_df = build_output_df(all_final, duplicates)
                xlsx_bytes = export_workbook(all_final, all_links)

                st.subheader("处理结果预览")
                st.dataframe(out_df, use_container_width=True)

                show_dashboard(out_df, BRAND_A, BRAND_B)

                st.download_button(
                    "下载结果Excel（含可点击链接）",
                    data=xlsx_bytes,
                    file_name=f"百度收录检测结果_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )

                # 存历史
                meta = {
                    "time": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "products": len(unique),
                    "duplicates": len(duplicates),
                    "mode": mode,
                    "runs": runs,
                    "good": int((out_df["收录效果状态"] == "好").sum()),
                    "bad": int((out_df["收录效果状态"] == "差").sum()),
                }
                stamp = save_history(out_df, meta)
                st.success(f"已存为历史记录：{stamp}")

    # ---------------- 历史记录页 ----------------
    with tab_history:
        st.subheader("历史记录")
        items = load_history_list()
        if not items:
            st.info("还没有历史记录。在「分析」页跑一次就会自动保存到这里。")
        else:
            for it in items:
                with st.expander(f"{it.get('time','')} ｜ {it.get('products',0)}个产品 ｜ {it.get('mode','')} ｜ 运行{it.get('runs',1)}次"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.write(f"好：{it.get('good',0)} ｜ 差：{it.get('bad',0)}")
                        st.write(f"重复：{it.get('duplicates',0)}")
                    with c2:
                        df = read_history_csv(it.get("stamp", ""))
                        if df is not None:
                            st.dataframe(df, use_container_width=True)
                            csv = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
                            st.download_button(
                                "下载此记录CSV",
                                data=csv,
                                file_name=f"{it.get('stamp','')}.csv",
                                mime="text/csv",
                            )


if __name__ == "__main__":
    main()
